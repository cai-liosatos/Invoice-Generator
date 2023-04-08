
import openpyxl
from openpyxl.styles import Alignment
import os.path
import win32com.client
from pywintypes import com_error
from views import dates_list, days_list

pay_information_dict = {
    "Weekday Support": ["15-045-0128-1-3", 45.0],
    "Saturday Support": ["01-013-0107-1-1", 70.0],
    "Sunday Support": ["01-014-0107-1-1", 75.0],
    "Public Holiday Support": ["01-012-0107-1-1", 90.0]
}
message = ''

# Functions
# recursion to find check files in output folder
def Pdf_check(dirlist):
    try:
        file = dirlist[-1].split('_')
    except:
        return False
    
    if (len(file) != 3) or (file[2][-3:] != 'pdf') or (file[0] != 'Invoice') or (len(file[2]) != 12):
        dirlist.pop(-1)
        file = Pdf_check(dirlist)
    if file:
        try:
            int(file[1][0:-2]) 
        except:
            dirlist.pop(-1)
            file = Pdf_check(dirlist)

    return file

# function to edit excel file
def Excel_edit(client, client_dict):    
    #  makes invoice_number either 1 greater than the last invoice, or 1 if none exist

    # Get list of all files only in the given directory
    dirlist = filter( lambda x: os.path.isfile(os.path.join(f'{fileDir}\Invoices', x)),
                            os.listdir(f'{fileDir}\Invoices') )
    # Sort list of files based on last modification time in ascending order
    dirlist = sorted( dirlist,
                            key = lambda x: os.path.getmtime(os.path.join(f'{fileDir}\Invoices', x))
                            )

    if len(dirlist) > 0:
        file = Pdf_check(dirlist)
        if file:
            invoice_number = int(file[1][0:-2]) + 1
        else:
            invoice_number = 1
        if message: return message
    else:
        invoice_number = 1

    # editing cells
    ws['C9'] = dates_list[-1]
    ws['C9'].alignment = ar
    ws['C10'] = "0"*(4-len(str(invoice_number))) + str(invoice_number)
    ws['C10'].alignment = ar
    ws['C12'] = client_dict["Name"][client_dict["person"]]
    ws['C12'].alignment = ar

    cell_number = 16
    for x in days_list:
        # identifying if worked with this client on a particular day
        if client_dict[x]["worked"][client_dict["person"]]:

            # identifying the type of shift (PH/weekend/weekday)
            if client_dict[x]["PH"][client_dict["person"]]:
                key = "Public Holiday Support"
            elif x == days_list[0]:
                key = "Sunday Support"
            elif x == days_list[-1]:
                key = "Saturday Support"
            else:
                key = "Weekday Support"
            
            # inputting correct information
            ws[f'A{cell_number}'] = dates_list[days_list.index(x)]
            ws[f'B{cell_number}'] = key
            ws[f'C{cell_number}'] = pay_information_dict[key][0]
            ws[f'D{cell_number}'] = client_dict[x]["Hours"][client_dict["person"]]
            ws[f'E{cell_number}'] = pay_information_dict[key][1]
            cell_number += 1

    # blanking out rest of price table
    if ws[f'A{cell_number}'].value: 
        for row in ws[f'A{cell_number}':'E25']:
            if not ws[f'A{cell_number}'].value:
                break
            cell_number += 1
            for cell in row:
                cell.value = None

    # populating KMs cell
    ws['D26'] = client_dict["KMs"][client_dict["person"]] if int(client_dict["KMs"][client_dict["person"]]) > 0 else None
           
    return invoice_number, xcl_file, message

# function to convert xcl file to pdf
def pdf_conversion(client, invoice_number, client_dict):
    # Converting to pdf
    WB_PATH = f'{fileDir}\{xcl_file}'
        
    # PDF path when saving
    PATH_TO_PDF = f'{fileDir}\Invoices\Invoice_{str(invoice_number)}{client_dict["Name"][client_dict["person"]].split(" ")[0][0]}{client_dict["Name"][client_dict["person"]].split(" ")[1][0]}_{dates_list[-1].replace("/","")}.pdf'
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    try:
        # Open
        wb = excel.Workbooks.open(WB_PATH)
        # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
        ws_index_list = [1]
        wb.WorkSheets(ws_index_list).Select()
        # Save
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error:
        return f"Sorry, we couldn't find the output file ({fileDir}\Invoices\). Is it possible this folder was moved, renamed or deleted?"
    else:
        wb.Close(True)
        excel.Quit()
        return ""

# master function, linking other functions together
def xc2pdf(clients, client_dict):
    for client in clients:
        client_dict["person"] = client_dict["Name"].index(client)
        if client in client_dict["Worked with"].keys():

            invoice_number, xcl_file, message = Excel_edit(client[0], client_dict)
            if message: return message

            wb.save(f"{fileDir}\{xcl_file}")
            message = pdf_conversion(client[0], invoice_number, client_dict)

    return "Successfully created invoices"

def main(client_dict):
    global fileDir, xcl_file, ar, wb, ws, message
    # Excel file path information
    fileDir = os.path.dirname(os.path.realpath('__file__'))
    xcl_file = 'Invoice-Template.xlsx'

    # Make invoice folder
    if not os.path.exists(f'{fileDir}\Invoices'):
        os.makedirs(f'{fileDir}\Invoices')

    # setting constant alignment variable
    ar = Alignment(horizontal='right')

    # importing worksheet
    try:
        wb = openpyxl.load_workbook(xcl_file)
        ws = wb.worksheets[0]
    except:
        message = f"Sorry, we couldn't find '{xcl_file}'. Is it possible it was moved, renamed or deleted?"
        
    if not message:
        message = xc2pdf([x for x in client_dict["Worked with"]], client_dict)
    return message