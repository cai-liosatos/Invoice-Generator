# close excel if error occurs between it opening and closing (cant find file error)
# mail setup?

import PyQt5 as pq
import sys
from PyQt5 import QtWidgets, uic, QtGui, QtCore
import openpyxl
from openpyxl.styles import Alignment
import datetime
import os.path
import win32com.client
from pywintypes import com_error
import ctypes

# dictionaries
map = {
    "Worked with": {},
    "person": 0,
    "Name": ["Elwyn Bourke", "Jonathan McNeill", "Zenith Burke"],
    "KMs": ["0", "0", 0],
    "Monday": {
        "worked": [True, True, False],
        "PH": [False, False, False],
        "Hours": ["1", "1", "0"]},
    "Tuesday": {
        "worked": [False, True, False],
        "PH": [False, False, False],
        "Hours": ["0", "1", "0"]},
    "Wednesday": {
        "worked": [True, True, False],
        "PH": [False, False, False],
        "Hours": ["1", "1", "0"]},
    "Thursday": {
        "worked": [True, True, False],
        "PH": [False, False, False],
        "Hours": ["1", "1", "0"]},
    "Friday": {
        "worked": [False, True, False],
        "PH": [False, False, False],
        "Hours": ["0", "1", "0"]},
    "Saturday": {
        "worked": [False, False, False],
        "PH": [False, False, False],
        "Hours": ["0", "0", "0"]},
    "Sunday": {
        "worked": [False, False, False],
        "PH": [False, False, False],
        "Hours": ["0", "0", "0"]}
}

pay_information_map = {
    "Weekday Support": ["15-045-0128-1-3", 45.0],
    "Saturday Support": ["01-013-0107-1-1", 70.0],
    "Sunday Support": ["01-014-0107-1-1", 75.0],
    "Public Holiday Support": ["01-012-0107-1-1", 90.0]
}

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# dynamically grabbing date of x days ago
def new_date(today, number):
    dt = datetime.timedelta(days = number)
    return today - dt

# GUI functionality

# function for confirming hours
def confirmation_setup():
    c_dlg.show()
    string = f"You are about to submit {len(map['Worked with'])} invoices with the following information:\n"
    for name in map["Worked with"]:
        string += f'\n{map["Worked with"][name][0]} Hour/s for {name} ({", ".join(map["Worked with"][name][1])})' 
    c_dlg.text_label.setText(string)

# main_view GUI functions
# function for updating text
def text_updating(labels, idxs=None, col=None):
    global dates_list
    global days_list
    # check if updating 'days' or default value in 'hours' input
    if idxs:
        for (label, day, idx) in zip(labels, days_list, idxs):
            label.setText(f'{day} ({dates_list[idx][:5]})') 
    else:
        for (label, day) in zip(labels, days_list):
            label.setText(map[day][col][map["person"]])

# function for updating isChecked of passed checkboxes
def checkbox_values(labels, col):
    global days_list
    for (label, day) in zip(labels, days_list):
        label.setChecked(map[day][col][map["person"]]) 

# master function for setting values when called
def setting_view():
    # heading label
    call.heading.setText(f'{dates_list[0]} -> {dates_list[-1]} - {map["Name"][map["person"]]}')

    # days col 
    text_updating(labels=[call.label_sun, call.label_mon, call.label_tues, call.label_wed, call.label_thurs, call.label_fri, call.label_sat], idxs=list(range(0,len(days_list)+1)))
    
    # Hours input
    text_updating(labels=[call.input_w_su, call.input_w_m, call.input_w_tu, call.input_w_w, call.input_w_th, call.input_w_f, call.input_w_sa], col="Hours")
    
    # Worked checkboxes
    checkbox_values([call.cb_w_su, call.cb_w_m, call.cb_w_tu, call.cb_w_w, call.cb_w_th, call.cb_w_f, call.cb_w_sa], "worked")
    
    # PH checkboxes
    checkbox_values([call.cb_ph_su, call.cb_ph_m, call.cb_ph_tu, call.cb_ph_w, call.cb_ph_th, call.cb_ph_f, call.cb_ph_sa], "PH")
    
    call.submitButton.setText("Next" if map["person"] < len(map["Name"]) - 1 else "Submit")
    call.prevButton.setVisible(map["person"] != 0)
    call.input_kms.setMaximum(999)

# function for updating map
def update_map(day, variables):
    if variables[0].isChecked():
        map[day]["worked"][map["person"]] = variables[0].isChecked()
        map[day]["PH"][map["person"]] = variables[1].isChecked()
        map[day]["Hours"][map["person"]] = variables[2].text()
        map["Worked with"][map["Name"][map["person"]]][0] += float(variables[2].text())
        map["Worked with"][map["Name"][map["person"]]][1].append(day[:2])

# function for checking validity of inputted values
def input_check(hours_labels, PH_labels):
    count = 0
    for l1, l2 in zip(hours_labels, PH_labels):
        try:
            float(l1.text())
        except ValueError:
            return "Please input either integer (e.g., 1) or float (e.g., 1.5) values into the hours column"
        count += 1 if l2.isChecked() else 0
    return 'There can only be a maximum of 3 public holiday shifts per client per week' if count > 3 else ''

# Functionality for prev button
def Previous():
    map["person"] -= 1
    if map["Name"][map["person"]] in map["Worked with"].keys():
        del map["Worked with"][map["Name"][map["person"]]]
    setting_view()

# functionality of skip button
def Next_client():
    map["person"] += 1
    if map["person"] < len(map["Name"]):
        setting_view()
        return
    confirmation_setup()

# functionality of submit button
def Submit(func=None):
    message = input_check([call.input_w_m, call.input_w_tu, call.input_w_w, call.input_w_th, call.input_w_f, call.input_w_sa, call.input_w_su],
                       [call.cb_ph_m, call.cb_ph_tu, call.cb_ph_w, call.cb_ph_th, call.cb_ph_f, call.cb_ph_sa, call.cb_ph_su])

    if message:
        ctypes.windll.user32.MessageBoxW(0, message, 1)
        return
    
    if func == 'skip':
        Next_client()
        return
    
    if map["person"] < len(map["Name"]):
    # update map with data, for each day 
        map["Worked with"][map["Name"][map["person"]]] = [0, []]
        update_map(days_list[0], [call.cb_w_su, call.cb_ph_su, call.input_w_su])
        update_map(days_list[1], [call.cb_w_m, call.cb_ph_m, call.input_w_m])
        update_map(days_list[2], [call.cb_w_tu, call.cb_ph_tu, call.input_w_tu])
        update_map(days_list[3], [call.cb_w_w, call.cb_ph_w, call.input_w_w])
        update_map(days_list[4], [call.cb_w_th, call.cb_ph_th, call.input_w_th])
        update_map(days_list[5], [call.cb_w_f, call.cb_ph_f, call.input_w_f])
        update_map(days_list[6], [call.cb_w_sa, call.cb_ph_sa, call.input_w_sa])

        if map["Worked with"][map["Name"][map["person"]]][0] == 0:
            del map["Worked with"][map["Name"][map["person"]]]
        # Update map with KMs
        map["KMs"][map["person"]] = call.input_kms.text()

        if map["person"] < len(map["Name"]) - 1:
            Next_client()
            return
    confirmation_setup()

# function for yes button on confirmation popup
def Dlg_Submit():
    global map_update
    map["person"] = 0
    c_dlg.close()
    call.close()
    map_update = True

# Identify date of end of week (today or last saturday if today isn't saturday)
date_sat = datetime.datetime.today()
date_idx = date_sat.weekday()
if date_idx != 5:
    date_sat = new_date(date_sat, date_idx - 5) if date_idx - 5 > 0 else new_date(date_sat, date_idx + 2)

# getting list of all dates in last week
dates_list = [new_date(date_sat, 6-x).strftime("%d/%m/%Y") for x in range(7)]
map_update = False
days_list = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

# Loading windows into memory from main_view.ui and confirmation_popup.ui
app=QtWidgets.QApplication([])
c_dlg=pq.uic.loadUi(resource_path("confirmation_popup.ui"))
call=pq.uic.loadUi(resource_path("main_view.ui"))

# Instantiating dynamic values and button functions
setting_view()
call.submitButton.clicked.connect(Submit)
call.prevButton.clicked.connect(Previous)
call.skipButton.clicked.connect(lambda: Submit('skip'))

c_dlg.noButton.clicked.connect(c_dlg.close)
c_dlg.yesButton.clicked.connect(Dlg_Submit)

# Opening window from memory
call.show()
app.exec_()

# quitting python file if main window is forced close
if not map_update: sys.exit()


# Excel and PDF Functionality

# Functions
# recursion to find check files in output folder
def Pdf_check(dirlist):
    try:
        file = dirlist[-1].split('_')
    except:
        return False
    
    if (len(file) != 3) or (file[2][-3:] != 'pdf') or (file[0] != 'Invoice') or (len(file[2]) != 12):
        dirlist.pop(-1)
        file = Pdf_check(dirlist)
    if file:
        try:
            int(file[1][0:-2]) 
        except:
            dirlist.pop(-1)
            file = Pdf_check(dirlist)

    return file

# function to edit excel file
def Excel_edit(client):
    # checking if folder exists, if not then creats it
    if not os.path.exists(f'{fileDir}\Invoices\{client}'):
        os.makedirs(f'{fileDir}\Invoices\{client}')
    
    #  makes invoice_number either 1 greater than the last invoice, or 1 if none exist

    # Get list of all files only in the given directory
    dirlist = filter( lambda x: os.path.isfile(os.path.join(f'{fileDir}\Invoices\{client}', x)),
                            os.listdir(f'{fileDir}\Invoices\{client}') )
    # Sort list of files based on last modification time in ascending order
    dirlist = sorted( dirlist,
                            key = lambda x: os.path.getmtime(os.path.join(f'{fileDir}\Invoices\{client}', x))
                            )

    if len(dirlist) > 0:
        file = Pdf_check(dirlist)
        if file:
            invoice_number = int(file[1][0:-2]) + 1
        else:
            invoice_number = 1
        if message: return message
    else:
        invoice_number = 1

    # editing cells
    ws['C9'] = dates_list[-1]
    ws['C9'].alignment = ar
    ws['C10'] = "0"*(4-len(str(invoice_number))) + str(invoice_number)
    ws['C10'].alignment = ar
    ws['C12'] = map["Name"][map["person"]]
    ws['C12'].alignment = ar

    cell_number = 16
    for x in days_list:
        # identifying if worked with this client on a particular day
        if map[x]["worked"][map["person"]]:

            # identifying the type of shift (PH/weekend/weekday)
            if map[x]["PH"][map["person"]]:
                key = "Public Holiday Support"
            elif x == days_list[0]:
                key = "Sunday Support"
            elif x == days_list[-1]:
                key = "Saturday Support"
            else:
                key = "Weekday Support"
            
            # inputting correct information
            ws[f'A{cell_number}'] = dates_list[days_list.index(x)]
            ws[f'B{cell_number}'] = key
            ws[f'C{cell_number}'] = pay_information_map[key][0]
            ws[f'D{cell_number}'] = map[x]["Hours"][map["person"]]
            ws[f'E{cell_number}'] = pay_information_map[key][1]
            cell_number += 1

    # blanking out rest of price table
    if ws[f'A{cell_number}'].value: 
        for row in ws[f'A{cell_number}':'E25']:
            if not ws[f'A{cell_number}'].value:
                break
            cell_number += 1
            for cell in row:
                cell.value = None

    # populating KMs cell
    ws['D26'] = map["KMs"][map["person"]] if int(map["KMs"][map["person"]]) > 0 else None
           
    return invoice_number, xcl_file, message

# function to convert xcl file to pdf
def pdf_conversion(client, invoice_number):
    # Converting to pdf
    WB_PATH = f'{fileDir}\{xcl_file}'
        
    # PDF path when saving
    PATH_TO_PDF = f'{fileDir}\Invoices\{client}\Invoice_{str(invoice_number)}{map["Name"][map["person"]].split(" ")[0][0]}{map["Name"][map["person"]].split(" ")[1][0]}_{dates_list[-1].replace("/","")}.pdf'
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    try:
        # Open
        wb = excel.Workbooks.open(WB_PATH)
        # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
        ws_index_list = [1]
        wb.WorkSheets(ws_index_list).Select()
        # Save
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error:
        return f"Sorry, we couldn't find the output file ({fileDir}\Invoices\{client}\). Is it possible this folder was moved, renamed or deleted?"
    else:
        wb.Close(True)
        excel.Quit()
        return ""

# master function, linking other functions together
def xc2pdf(clients):
    for client in clients:
        map["person"] = map["Name"].index(client)
        if client in map["Worked with"].keys():

            invoice_number, xcl_file, message = Excel_edit(client[0])
            if message: return message

            wb.save(f"{fileDir}\{xcl_file}")
            message = pdf_conversion(client[0], invoice_number)

    return "Successfully created invoices"

message = ''
# Excel file path information
fileDir = os.path.dirname(os.path.realpath('__file__'))
xcl_file = 'Invoice-Template.xlsx'

# Make invoice folder
if not os.path.exists(f'{fileDir}\Invoices'):
    os.makedirs(f'{fileDir}\Invoices')

# setting constant alignment variable
ar = Alignment(horizontal='right')

# importing worksheet
try:
    wb = openpyxl.load_workbook(xcl_file)
    ws = wb.worksheets[0]
except:
    message = f"Sorry, we couldn't find '{xcl_file}'. Is it possible it was moved, renamed or deleted?"
    
if not message:
    message = xc2pdf([x for x in map["Worked with"]])
ctypes.windll.user32.MessageBoxW(0, message, 1)